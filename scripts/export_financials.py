#!/usr/bin/env python3
"""
export_financials.py

Reads AMC_Financial_Statements.xlsx and outputs site/data/financials.json.
Falls back to hardcoded values if the Excel file cannot be read.

LTM formula: FY2024 + 9M2025 - 9M2024
"""

import json
import os
import sys

EXCEL_PATH = (
    "/Users/diego/school/binghamton-non-school/BRx/Spring26CaseComp/"
    "Case Competition Materials (AMC) 2/AMC_Financial_Statements.xlsx"
)
OUTPUT_PATH = (
    "/Users/diego/school/binghamton-non-school/BRx/Spring26CaseComp/"
    "site/data/financials.json"
)

# ---------------------------------------------------------------------------
# Hardcoded fallback data (sourced directly from the Excel extraction)
# Columns: FY2022, FY2023, FY2024, 9M2024, 9M2025, Q32024, Q32025
# ---------------------------------------------------------------------------

HARDCODED = {
    "asOfDate": "2025-09-30",
    "periods": ["FY 2022", "FY 2023", "FY 2024", "9M 2024", "9M 2025", "Q3 2024", "Q3 2025"],

    # --- Income Statement ---
    "incomeStatement": {
        "admissions":       [2201.4, 2690.5, 2560.5, 1839.1, 1951.2, 744.2, 715.1],
        "fnb":              [1313.7, 1669.8, 1624.9, 1178.7, 1234.8, 490.4, 451.8],
        "other":            [396.3,  452.3,  451.8,  313.0,  374.6,  114.2, 133.3],
        "revenue":          [3911.4, 4812.6, 4637.2, 3330.8, 3560.6, 1348.8, 1300.2],
        "filmCost":         [1051.7, 1291.1, 1239.2, 893.0,  949.3,  381.4, 352.4],
        "fnbCost":          [228.6,  315.3,  305.6,  222.6,  241.9,  89.7,  88.6],
        "opex":             [1528.4, 1691.5, 1679.4, 1237.9, 1316.3, 454.6, 464.7],
        "rent":             [886.2,  873.5,  873.6,  659.3,  664.8,  216.4, 224.1],
        "gaMa":             [2.1,    1.7,    0.1,    0.1,    3.2,    0.1,   0.1],
        "gaOther":          [207.6,  241.9,  226.8,  160.7,  169.3,  54.0,  55.1],
        "da":               [396.0,  365.0,  319.5,  241.2,  233.3,  80.8,  79.4],
        "impairment":       [133.1,  106.9,  72.3,   0.0,    0.0,    0.0,   0.0],
        "ebitda":           [6.8,    397.6,  312.5,  157.2,  215.8,  152.6, 115.2],
        "ebitdaMargin":     [0.00174, 0.08262, 0.06739, 0.04720, 0.06061, 0.11314, 0.08860],
        "otherExpense":     [55.2,  -76.8,  -156.2, -173.8, 103.9,  -22.8, 194.8],
        "interestCorp":     [336.4,  369.6,  401.8,  289.8,  337.6,  109.6, 119.0],
        "interestLease":    [4.1,    3.7,    5.4,    2.5,    4.3,    1.0,   1.7],
        "ncm":              [38.2,   37.9,   36.5,   27.5,   46.1,   9.0,   18.6],
        "investmentIncome": [14.9,  -15.5,  -16.3,  -14.4,  -8.4,  -3.2,  -1.3],
        "netLoss":          [-973.6, -396.6, -352.6, -217.0, -505.0, -20.7, -298.2],
        "eps":              [-9.29,  -2.37,  -1.06,  -0.69,  -1.10,  -0.06, -0.58],
        "sharesOutstanding": [104769, 167644, 332920, 315783, 459375, 361853, 513010],
    },

    # --- Balance Sheet ---
    # Columns: Dec 31 2023, Dec 31 2024, Sep 30 2025
    "balanceSheet": {
        "periods": ["Dec 31, 2023", "Dec 31, 2024", "Sep 30, 2025"],
        "cash":                  [884.3,  632.3,  365.8],
        "restrictedCash":        [27.1,   48.5,   51.1],
        "receivables":           [203.7,  168.1,  102.3],
        "otherCurrent":          [88.0,   98.3,   99.8],
        "totalCurrentAssets":    [1203.1, 947.2,  619.0],
        "propertyNet":           [1560.4, 1442.3, 1410.8],
        "rouAssets":             [3544.5, 3220.1, 3230.9],
        "intangibles":           [146.7,  144.3,  147.6],
        "goodwill":              [2358.7, 2301.1, 2400.0],
        "otherLongTerm":         [195.8,  192.5,  212.4],
        "totalAssets":           [9009.2, 8247.5, 8020.7],
        "accountsPayable":       [320.5,  378.3,  279.1],
        "accruedExpenses":       [350.8,  340.6,  334.9],
        "deferredRevenue":       [421.8,  432.4,  411.5],
        "currentDebt":           [25.1,   64.2,   19.9],
        "currentFinanceLease":   [5.4,    4.4,    5.5],
        "currentOpLease":        [508.8,  524.9,  552.5],
        "totalCurrentLiabilities": [1632.4, 1744.8, 1603.4],
        "corporateBorrowings":   [4552.3, 4010.9, 3990.1],
        "financeLease":          [50.0,   44.9,   47.4],
        "opLease":               [4000.7, 3627.6, 3581.5],
        "exhibitorAgreement":    [486.6,  464.0,  460.5],
        "otherLongTermLiab":     [135.1,  115.8,  115.3],
        "totalLiabilities":      [10857.1, 10008.0, 9798.2],
        "accumulatedDeficit":    [-7994.2, -8346.8, -8851.8],
        "stockholdersDeficit":   [-1847.9, -1760.5, -1777.5],
        "totalDebt":             [4577.4, 4075.1, 4010.0],
        "netDebt":               [3693.1, 3442.8, 3644.2],
    },

    # --- Cash Flow ---
    # Columns: FY2022, FY2023, FY2024, 9M2024, 9M2025
    "cashFlow": {
        "periods": ["FY 2022", "FY 2023", "FY 2024", "9M 2024", "9M 2025"],
        "netLoss":            [-973.6, -396.6, -352.6, -217.0, -505.0],
        "da":                 [396.0,  365.0,  319.5,  241.2,  233.3],
        "debtExtinguishment": [92.8,  -142.8,  -38.9,  -40.3,  196.0],
        "pikInterest":        [0.0,    0.0,    14.7,   0.0,    28.7],
        "impairment":         [133.1,  106.9,  72.3,   0.0,    0.0],
        "cfOperating":        [-628.5, -215.2, -50.8,  -254.4, -246.5],
        "capex":              [-202.0, -225.6, -245.5, -155.8, -162.7],
        "cfInvesting":        [-224.0, -180.1, -242.9, -154.0, -163.9],
        "equityIssuances":    [220.4,  832.7,  254.9,  243.0,  169.6],
        "debtProceeds":       [1318.0, 0.0,    27.0,   27.0,   244.4],
        "debtRepayments":     [-1509.5, -161.6, -160.1, -138.2, -234.5],
        "cfFinancing":        [-91.3,  649.3,  68.4,   72.1,   134.3],
        "netCashChange":      [-965.9, 257.0,  -230.6, -334.3, -263.9],
        "endingCash":         [654.4,  911.4,  680.8,  577.1,  416.9],
    },
}

# Index positions within the 7-column income statement arrays
IDX_FY2024  = 2   # FY 2024
IDX_9M2024  = 3   # 9M 2024
IDX_9M2025  = 4   # 9M 2025

# Index positions within the 5-column cash flow arrays
CF_IDX_FY2024 = 2
CF_IDX_9M2024 = 3
CF_IDX_9M2025 = 4


def compute_ltm(series, idx_fy24=IDX_FY2024, idx_9m24=IDX_9M2024, idx_9m25=IDX_9M2025):
    """Return LTM = FY2024 + 9M2025 - 9M2024, rounded to 1 decimal."""
    return round(series[idx_fy24] + series[idx_9m25] - series[idx_9m24], 1)


def build_from_hardcoded():
    """Assemble the output dict from the hardcoded data and compute LTM."""
    IS = HARDCODED["incomeStatement"]
    data = {
        "asOfDate": HARDCODED["asOfDate"],
        "periods": HARDCODED["periods"],
        "ltm": {
            "revenue":      compute_ltm(IS["revenue"]),
            "ebitda":       compute_ltm(IS["ebitda"]),
            "interestCorp": compute_ltm(IS["interestCorp"]),
            "netLoss":      compute_ltm(IS["netLoss"]),
        },
        "incomeStatement": {
            "periods": HARDCODED["periods"],
            **IS,
        },
        "balanceSheet": HARDCODED["balanceSheet"],
        "cashFlow": HARDCODED["cashFlow"],
    }
    return data


# ---------------------------------------------------------------------------
# Excel reader (optional — used when the file is available)
# ---------------------------------------------------------------------------

def safe_float(val):
    """Convert a cell value to float, returning 0.0 on failure."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def read_row(ws, row_num, col_start, col_end):
    """Read a row slice from the worksheet, returning a list of floats."""
    return [safe_float(ws.cell(row=row_num, column=c).value)
            for c in range(col_start, col_end + 1)]


def try_read_excel():
    """
    Attempt to read the Excel file with openpyxl.
    Returns a dict in the same shape as build_from_hardcoded(), or None on error.

    NOTE: This function maps expected row labels to row numbers based on the
    known layout of AMC_Financial_Statements.xlsx. If the sheet structure
    changes, update ROW_MAP below.
    """
    try:
        import openpyxl
    except ImportError:
        print("[WARN] openpyxl not installed — using hardcoded data.", file=sys.stderr)
        return None

    if not os.path.exists(EXCEL_PATH):
        print(f"[WARN] Excel file not found: {EXCEL_PATH}", file=sys.stderr)
        print("[WARN] Using hardcoded data.", file=sys.stderr)
        return None

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except Exception as exc:
        print(f"[WARN] Could not open Excel: {exc}", file=sys.stderr)
        return None

    # -----------------------------------------------------------------------
    # Sheet: Income Statement
    # Expected column layout (1-indexed):
    #   Col B=FY2022, C=FY2023, D=FY2024, E=9M2024, F=9M2025, G=Q32024, H=Q32025
    # -----------------------------------------------------------------------
    IS_COL_START = 2   # column B
    IS_COL_END   = 8   # column H  (7 columns total)

    # Row numbers in the Income Statement sheet (adjust if layout differs)
    IS_ROW = {
        "admissions":       4,
        "fnb":              5,
        "other":            6,
        "revenue":          7,
        "filmCost":         9,
        "fnbCost":          10,
        "opex":             11,
        "rent":             12,
        "gaMa":             13,
        "gaOther":          14,
        "da":               15,
        "impairment":       16,
        "ebitda":           18,
        "ebitdaMargin":     19,
        "otherExpense":     21,
        "interestCorp":     22,
        "interestLease":    23,
        "ncm":              24,
        "investmentIncome": 25,
        "netLoss":          27,
        "eps":              28,
        "sharesOutstanding": 29,
    }

    # -----------------------------------------------------------------------
    # Sheet: Balance Sheet
    # Expected column layout (1-indexed):
    #   Col B=Dec31_2023, C=Dec31_2024, D=Sep30_2025
    # -----------------------------------------------------------------------
    BS_COL_START = 2
    BS_COL_END   = 4

    BS_ROW = {
        "cash":                  4,
        "restrictedCash":        5,
        "receivables":           6,
        "otherCurrent":          7,
        "totalCurrentAssets":    8,
        "propertyNet":           10,
        "rouAssets":             11,
        "intangibles":           12,
        "goodwill":              13,
        "otherLongTerm":         14,
        "totalAssets":           15,
        "accountsPayable":       17,
        "accruedExpenses":       18,
        "deferredRevenue":       19,
        "currentDebt":           20,
        "currentFinanceLease":   21,
        "currentOpLease":        22,
        "totalCurrentLiabilities": 23,
        "corporateBorrowings":   25,
        "financeLease":          26,
        "opLease":               27,
        "exhibitorAgreement":    28,
        "otherLongTermLiab":     29,
        "totalLiabilities":      30,
        "accumulatedDeficit":    32,
        "stockholdersDeficit":   33,
        "totalDebt":             35,
        "netDebt":               36,
    }

    # -----------------------------------------------------------------------
    # Sheet: Cash Flow
    # Expected column layout (1-indexed):
    #   Col B=FY2022, C=FY2023, D=FY2024, E=9M2024, F=9M2025
    # -----------------------------------------------------------------------
    CF_COL_START = 2
    CF_COL_END   = 6

    CF_ROW = {
        "netLoss":            4,
        "da":                 5,
        "debtExtinguishment": 6,
        "pikInterest":        7,
        "impairment":         8,
        "cfOperating":        10,
        "capex":              12,
        "cfInvesting":        13,
        "equityIssuances":    15,
        "debtProceeds":       16,
        "debtRepayments":     17,
        "cfFinancing":        18,
        "netCashChange":      20,
        "endingCash":         21,
    }

    # Helper: try to get a worksheet by name, falling back to index
    def get_sheet(names_to_try, fallback_index):
        for name in names_to_try:
            if name in wb.sheetnames:
                return wb[name]
        if fallback_index < len(wb.sheetnames):
            return wb[wb.sheetnames[fallback_index]]
        return None

    ws_is = get_sheet(["Income Statement", "IS", "IncomeStatement"], 0)
    ws_bs = get_sheet(["Balance Sheet", "BS", "BalanceSheet"], 1)
    ws_cf = get_sheet(["Cash Flow", "CF", "CashFlow"], 2)

    if ws_is is None or ws_bs is None or ws_cf is None:
        print("[WARN] Could not identify all three sheets — using hardcoded data.",
              file=sys.stderr)
        return None

    # Read Income Statement
    IS = {}
    for key, row in IS_ROW.items():
        IS[key] = read_row(ws_is, row, IS_COL_START, IS_COL_END)

    # Read Balance Sheet
    BS = {}
    BS["periods"] = ["Dec 31, 2023", "Dec 31, 2024", "Sep 30, 2025"]
    for key, row in BS_ROW.items():
        BS[key] = read_row(ws_bs, row, BS_COL_START, BS_COL_END)

    # Read Cash Flow
    CF = {}
    CF["periods"] = ["FY 2022", "FY 2023", "FY 2024", "9M 2024", "9M 2025"]
    for key, row in CF_ROW.items():
        CF[key] = read_row(ws_cf, row, CF_COL_START, CF_COL_END)

    periods = ["FY 2022", "FY 2023", "FY 2024", "9M 2024", "9M 2025", "Q3 2024", "Q3 2025"]

    # Compute LTM (indices 2, 3, 4 within IS arrays)
    ltm = {
        "revenue":      compute_ltm(IS["revenue"]),
        "ebitda":       compute_ltm(IS["ebitda"]),
        "interestCorp": compute_ltm(IS["interestCorp"]),
        "netLoss":      compute_ltm(IS["netLoss"]),
    }

    return {
        "asOfDate": "2025-09-30",
        "periods":  periods,
        "ltm":      ltm,
        "incomeStatement": {"periods": periods, **IS},
        "balanceSheet":    BS,
        "cashFlow":        CF,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("AMC Financial Data Exporter")
    print("=" * 40)

    # Try Excel first; fall back to hardcoded
    data = try_read_excel()
    if data is not None:
        print(f"[OK] Data read from Excel: {EXCEL_PATH}")
        # Sanity-check: if revenue looks wildly wrong, fall back
        try:
            rev_fy24 = data["incomeStatement"]["revenue"][IDX_FY2024]
            if not (1000 < rev_fy24 < 10000):
                print("[WARN] Excel revenue value looks suspect — using hardcoded data.",
                      file=sys.stderr)
                data = build_from_hardcoded()
        except Exception:
            data = build_from_hardcoded()
    else:
        data = build_from_hardcoded()
        print("[OK] Using hardcoded fallback data.")

    # Round all numeric values to 1 decimal (except shares and margins)
    # Shares are integers; margins are already stored as ratios with 5 decimals.
    # We leave those as-is and only round the monetary figures.
    def round_series(lst, decimals=1):  # noqa: E301
        return [round(v, decimals) if isinstance(v, float) else v for v in lst]

    for section in ("incomeStatement", "cashFlow"):
        for key, val in data[section].items():
            if isinstance(val, list) and key not in ("periods", "sharesOutstanding", "ebitdaMargin", "eps"):
                data[section][key] = round_series(val)
            elif isinstance(val, list) and key == "eps":
                # EPS reported to 2 decimal places
                data[section][key] = round_series(val, decimals=2)

    for key, val in data["balanceSheet"].items():
        if isinstance(val, list) and key != "periods":
            data["balanceSheet"][key] = round_series(val)

    for key, val in data["ltm"].items():
        if isinstance(val, (int, float)):
            data["ltm"][key] = round(val, 1)

    # Write output
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)

    print(f"[OK] Written to: {OUTPUT_PATH}")
    print()

    # Quick verification summary
    ltm = data["ltm"]
    print("LTM Summary:")
    print(f"  Revenue:       ${ltm['revenue']:.1f}M")
    print(f"  EBITDA:        ${ltm['ebitda']:.1f}M")
    print(f"  Interest Corp: ${ltm['interestCorp']:.1f}M")
    print(f"  Net Loss:      ${ltm['netLoss']:.1f}M")

    bs = data["balanceSheet"]
    print()
    print("Balance Sheet (most recent — Sep 30, 2025):")
    print(f"  Cash:          ${bs['cash'][-1]:.1f}M")
    print(f"  Total Debt:    ${bs['totalDebt'][-1]:.1f}M")
    print(f"  Net Debt:      ${bs['netDebt'][-1]:.1f}M")

    return 0


if __name__ == "__main__":
    sys.exit(main())
